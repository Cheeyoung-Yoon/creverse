# %%
import asyncio
import json
import logging
import pandas as pd
import requests
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('batch_evaluation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EssayBatchEvaluator:
    """배치 에세이 평가기 - 다중 prompt 버전 지원, 중간 저장 기능 포함"""
    
    def __init__(self, api_url: str = "http://localhost:8000/v1/essay-eval", prompt_versions: List[str] = None, checkpoint_file: str = None):
        self.api_url = api_url
        self.levels = ["Basic", "Intermediate", "Advanced", "Expert"]
        self.prompt_versions = prompt_versions or ["v1.2.0", "v1.4.1"]
        self.results = {}
        self.checkpoint_file = checkpoint_file or "batch_evaluation_checkpoint.json"
        self.batch_size = 5  # 5개 API 호출마다 저장
        self.progress = {"completed_calls": 0, "total_calls": 0, "current_position": None}
        
        logger.info(f"🔧 Initialized evaluator with prompt versions: {self.prompt_versions}")
        logger.info(f"📊 Total combinations: {len(self.levels)} levels × {len(self.prompt_versions)} versions = {len(self.levels) * len(self.prompt_versions)} per essay")
        logger.info(f"💾 Checkpoint file: {self.checkpoint_file}")
        logger.info(f"📦 Batch size: {self.batch_size} calls per save")
        
    def save_checkpoint(self):
        """현재 진행 상황을 JSON 파일로 저장"""
        checkpoint_data = {
            "progress": self.progress,
            "results": self.results,
            "config": {
                "prompt_versions": self.prompt_versions,
                "levels": self.levels,
                "api_url": self.api_url
            },
            "timestamp": datetime.now().isoformat()
        }
        
        try:
            with open(self.checkpoint_file, 'w', encoding='utf-8') as f:
                json.dump(checkpoint_data, f, ensure_ascii=False, indent=2)
            logger.debug(f"💾 Checkpoint saved: {self.progress['completed_calls']}/{self.progress['total_calls']} calls")
        except Exception as e:
            logger.error(f"❌ Failed to save checkpoint: {e}")
    
    def load_checkpoint(self) -> bool:
        """저장된 checkpoint가 있으면 로드"""
        try:
            if not Path(self.checkpoint_file).exists():
                logger.info("🔄 No existing checkpoint found, starting fresh")
                return False
                
            with open(self.checkpoint_file, 'r', encoding='utf-8') as f:
                checkpoint_data = json.load(f)
            
            self.progress = checkpoint_data.get("progress", {})
            self.results = checkpoint_data.get("results", {})
            
            logger.info(f"✅ Checkpoint loaded: {self.progress.get('completed_calls', 0)}/{self.progress.get('total_calls', 0)} calls completed")
            logger.info(f"📅 Checkpoint timestamp: {checkpoint_data.get('timestamp', 'unknown')}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to load checkpoint: {e}")
            return False
    
    def should_skip_call(self, level: str, version: str, essay_idx: int) -> bool:
        """이미 완료된 호출인지 확인"""
        key = f"{level}_{version}"
        if key not in self.results:
            return False
            
        # 해당 에세이가 이미 처리되었는지 확인
        existing_results = self.results[key]
        for result in existing_results:
            if result.get("essay_id") == essay_idx or result.get("essay_index") == essay_idx:
                return True
        return False
        
    def load_sample_data(self, excel_path: str) -> pd.DataFrame:
        """샘플 에세이 데이터 로드"""
        try:
            df = pd.read_excel(excel_path)
            # 첫 번째 행이 헤더인지 확인하고 NaN 값이 있는 행 제거
            df = df.dropna(subset=['submit_text'])
            
            logger.info(f"✅ Loaded {len(df)} essays from {excel_path}")
            logger.info(f"📊 Columns: {list(df.columns)}")
            logger.info(f"📋 Original levels distribution:")
            if 'rubric_level' in df.columns:
                level_counts = df['rubric_level'].value_counts()
                for level, count in level_counts.items():
                    logger.info(f"   {level}: {count} essays")
            
            return df
        except Exception as e:
            logger.error(f"❌ Failed to load Excel file: {e}")
            raise
    
    def call_evaluation_api(self, essay_text: str, topic_prompt: str, level_group: str, prompt_version: str = "v1.4.1") -> Dict[str, Any]:
        """API 호출하여 에세이 평가"""
        payload = {
            "rubric_level": level_group,
            "topic_prompt": topic_prompt,
            "submit_text": essay_text,
            "prompt_version": prompt_version
        }
        
        try:
            response = requests.post(
                self.api_url,
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=60  # 60초 타임아웃
            )
            
            if response.status_code == 200:
                result = response.json()
                logger.debug(f"✅ API call successful for level {level_group}")
                return {
                    "status": "success",
                    "data": result,
                    "response_time": response.elapsed.total_seconds()
                }
            else:
                logger.error(f"❌ API call failed with status {response.status_code}")
                return {
                    "status": "error",
                    "error": f"HTTP {response.status_code}: {response.text}",
                    "response_time": response.elapsed.total_seconds()
                }
                
        except requests.exceptions.Timeout:
            logger.error("❌ API call timed out")
            return {"status": "timeout", "error": "Request timed out"}
        except Exception as e:
            logger.error(f"❌ API call failed: {e}")
            return {"status": "error", "error": str(e)}
    
    def process_all_essays(self, df: pd.DataFrame) -> Dict[str, List[Dict]]:
        """모든 에세이를 모든 레벨과 prompt 버전으로 평가 (checkpoint 지원)"""
        
        # checkpoint 로드 시도
        checkpoint_loaded = self.load_checkpoint()
        
        # 결과 딕셔너리 초기화 (checkpoint에서 로드되지 않은 경우)
        if not checkpoint_loaded:
            self.results = {}
            for level in self.levels:
                for version in self.prompt_versions:
                    key = f"{level}_{version}"
                    self.results[key] = []
        else:
            # checkpoint에서 로드된 경우, 누락된 키들을 초기화
            for level in self.levels:
                for version in self.prompt_versions:
                    key = f"{level}_{version}"
                    if key not in self.results:
                        self.results[key] = []
                        logger.info(f"🔧 Initialized missing key: {key}")
        
        # 현재 결과 상태 로그
        logger.info("📊 Current results status:")
        for key, results_list in self.results.items():
            logger.info(f"  {key}: {len(results_list)} completed essays")
        
        total_calls = len(df) * len(self.levels) * len(self.prompt_versions)
        current_call = self.progress.get("completed_calls", 0)
        
        # 진행 상황 업데이트
        self.progress["total_calls"] = total_calls
        
        logger.info(f"🚀 Starting batch evaluation: {len(df)} essays × {len(self.levels)} levels × {len(self.prompt_versions)} versions = {total_calls} API calls")
        if checkpoint_loaded:
            logger.info(f"🔄 Resuming from checkpoint: {current_call}/{total_calls} calls already completed")
        
        # 초기 체크포인트 저장 (시작 시점)
        self.save_checkpoint()
        logger.info(f"💾 Initial checkpoint saved")
        
        try:
            for level in self.levels:
                for version in self.prompt_versions:
                    logger.info(f"📊 Processing level: {level}, version: {version}")
                    
                    for idx, row in df.iterrows():
                        # 이미 완료된 호출인지 확인
                        if self.should_skip_call(level, version, idx):
                            logger.debug(f"⏭️ Skipping already completed: essay {idx}, level {level}, version {version}")
                            continue
                        
                        current_call += 1
                    essay_text = str(row.get('submit_text', ''))
                    essay_id = row.get('essay_id', idx)
                    original_level = row.get('rubric_level', 'unknown')
                    topic_prompt = row.get('topic_prompt', '')
                    
                    if not essay_text or essay_text.strip() == '':
                        logger.warning(f"⚠️ Empty essay text at row {idx} (essay_id: {essay_id})")
                        continue
                    
                    logger.info(f"[{current_call}/{total_calls}] Evaluating essay {essay_id} (original: {original_level}) with level {level}, version {version}")
                    logger.debug(f"📝 Essay preview: {essay_text[:100]}...")
                    
                    # 진행 상황 업데이트
                    self.progress["completed_calls"] = current_call
                    self.progress["current_position"] = {
                        "level": level,
                        "version": version,
                        "essay_idx": idx,
                        "essay_id": essay_id
                    }
                    
                    # API 호출
                    api_result = self.call_evaluation_api(essay_text, topic_prompt, level, version)
                    
                    # 결과 정리
                    # 기본 정보 기록
                    result_record = {
                        "essay_id": essay_id,
                        "original_level": original_level,
                        "evaluation_level": level,
                        "prompt_version": version,  # prompt 버전 정보 추가
                        "topic_prompt": topic_prompt,
                        "essay_text": essay_text[:500] + "..." if len(essay_text) > 500 else essay_text,  # 텍스트 길이 제한
                        "essay_length": len(essay_text),
                        "response_time": api_result.get("response_time", 0),
                        "api_status": api_result.get("status", "unknown")
                    }
                
                if api_result["status"] == "success":
                    eval_data = api_result["data"]
                    
                    # 전체 평가 정보 추가
                    # if "aggregated" in eval_data:
                    #     agg_data = eval_data["aggregated"]
                    #     result_record["total_score"] = agg_data.get("score", 0)
                    #     result_record["total_corrections_count"] = len(agg_data.get("corrections", []))
                    
                    # grammar 섹션 처리
                    if "grammar" in eval_data:
                        grammar = eval_data["grammar"]
                        result_record["grammar_score"] = grammar.get("score", 0)
                        result_record["grammar_feedback"] = grammar.get("feedback", "")[:500]
                        result_record["grammar_corrections_count"] = len(grammar.get("corrections", []))
                        
                        corrections = grammar.get("corrections", [])
                        if corrections:
                            first_correction = corrections[0]
                            result_record["grammar_first_correction"] = f"{first_correction.get('highlight', '')} → {first_correction.get('correction', '')}"[:200]
                    
                    # structure 안의 섹션들 처리
                    if "structure" in eval_data:
                        structure_data = eval_data["structure"]
                        for section_name in ["introduction", "body", "conclusion"]:
                            if section_name in structure_data:
                                section = structure_data[section_name]
                                result_record[f"{section_name}_score"] = section.get("score", 0)
                                result_record[f"{section_name}_feedback"] = section.get("feedback", "")[:500]
                                result_record[f"{section_name}_corrections_count"] = len(section.get("corrections", []))
                                
                                # 첫 번째 correction만 기록
                                corrections = section.get("corrections", [])
                                if corrections:
                                    first_correction = corrections[0]
                                    result_record[f"{section_name}_first_correction"] = f"{first_correction.get('highlight', '')} → {first_correction.get('correction', '')}"[:200]
                    
                    # 타이밍 정보
                    if "timings" in eval_data:
                        result_record["total_processing_time"] = eval_data["timings"].get("total", 0) / 1000  # ms를 초로 변환
                        
                else:
                    result_record["error"] = api_result.get("error", "Unknown error")
                
                # 레벨과 버전 조합 키로 결과 저장
                key = f"{level}_{version}"
                result_record["essay_index"] = idx  # checkpoint에서 사용할 인덱스 추가
                
                # 키가 존재하는지 확인하고 초기화
                if key not in self.results:
                    logger.warning(f"⚠️ Key {key} not found in results, initializing")
                    self.results[key] = []
                
                self.results[key].append(result_record)
                logger.debug(f"📝 Added result for {key}, total count: {len(self.results[key])}")
                
                # 배치 단위로 checkpoint 저장
                if current_call % self.batch_size == 0:
                    self.save_checkpoint()
                    logger.info(f"💾 Checkpoint saved at {current_call}/{total_calls} calls")
                    # 현재 결과 개수 로그
                    for k, v in self.results.items():
                        logger.info(f"  {k}: {len(v)} results")
                
                        # API 호출 간격 (서버 부하 방지)
                        time.sleep(1)
        
        except KeyboardInterrupt:
            logger.info("\n⏹️ Process interrupted by user")
            self.save_checkpoint()
            logger.info(f"💾 Progress saved in checkpoint: {self.checkpoint_file}")
            raise
        
        # 최종 checkpoint 저장
        self.save_checkpoint()
        logger.info("✅ Batch evaluation completed!")
        return self.results
    
    def create_excel_report(self, results: Dict[str, List[Dict]], output_path: str):
        """결과를 엑셀 파일로 저장 - 레벨별, 버전별 시트 생성"""
        logger.info(f"📝 Creating Excel report: {output_path}")
        
        # 디버깅: 사용 가능한 키들 출력
        logger.info(f"🔍 Available result keys: {list(results.keys())}")
        for key, value in results.items():
            logger.info(f"  {key}: {len(value)} results")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 각 레벨과 버전 조합에 대한 시트 생성
            for level in self.levels:
                for version in self.prompt_versions:
                    # 키 검색 시도 (여러 형식으로)
                    possible_keys = [
                        f"{level}_{version}",  # 원본 형식 (예: Basic_v1.5.0)
                        f"{level}_{version.replace('.', '_')}",  # 점을 언더스코어로 (예: Basic_v1_5_0)
                    ]
                    
                    level_version_results = []
                    found_key = None
                    
                    for key in possible_keys:
                        if key in results:
                            level_version_results = results[key]
                            found_key = key
                            logger.info(f"✓ Found data for key: {key}")
                            break
                    
                    if not level_version_results:
                        logger.warning(f"⚠️ No results for level {level}, version {version}")
                        logger.warning(f"   Tried keys: {possible_keys}")
                        continue
                    
                    # DataFrame 생성
                    df = pd.DataFrame(level_version_results)
                    
                    # 시트 이름 설정 (Excel 시트명 길이 제한 고려, 점을 언더스코어로 변경)
                    sheet_name = f"{level}_{version.replace('.', '_')}"[:31]  # Excel 시트명 최대 31자
                    
                    logger.info(f"📊 Creating sheet '{sheet_name}' using key '{found_key}' with {len(level_version_results)} results")
                    
                    # 시트에 저장
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # 스타일링
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # 헤더 스타일링
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    
                    # 열 너비 자동 조정
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # 최대 50자
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                    logger.info(f"✅ Created sheet: {sheet_name} with {len(level_version_results)} records")
            
            # 요약 시트 생성
            self.create_summary_sheet(writer, results)
        
        logger.info(f"✅ Excel report saved: {output_path}")
    
    def create_summary_sheet(self, writer, results: Dict[str, List[Dict]]):
        """요약 시트 생성 - 레벨별, 버전별 통계"""
        summary_data = []
        
        for level in self.levels:
            for version in self.prompt_versions:
                key = f"{level}_{version}"
                level_version_results = results.get(key, [])
                
                if not level_version_results:
                    continue
                
                successful_calls = len([r for r in level_version_results if r.get("api_status") == "success"])
                failed_calls = len([r for r in level_version_results if r.get("api_status") != "success"])
                avg_response_time = sum([r.get("response_time", 0) for r in level_version_results]) / len(level_version_results)
                
                # 평균 점수 계산 (성공한 호출만)
                successful_results = [r for r in level_version_results if r.get("api_status") == "success"]
                avg_scores = {}
                
                if successful_results:
                    sections = ["introduction", "body", "conclusion", "grammar"]
                    for section in sections:
                        scores = [r.get(f"{section}_score", 0) for r in successful_results if f"{section}_score" in r]
                        avg_scores[f"avg_{section}_score"] = sum(scores) / len(scores) if scores else 0
                
                summary_record = {
                    "Level": level,
                    "Prompt_Version": version,
                    "Total_Essays": len(level_version_results),
                    "Successful_Calls": successful_calls,
                    "Failed_Calls": failed_calls,
                    "Success_Rate": f"{(successful_calls/len(level_version_results)*100):.1f}%" if level_version_results else "0%",
                    "Avg_Response_Time": f"{avg_response_time:.2f}s",
                    **avg_scores
                }
                
                summary_data.append(summary_record)
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        
        # 요약 시트 스타일링
        workbook = writer.book
        worksheet = writer.sheets["Summary"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        logger.info("✅ Created Summary sheet")

def main():
    """메인 실행 함수"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Essay Batch Evaluation with Multiple Prompt Versions")
    parser.add_argument("--versions", 
                       default="v1.5.0",
                       help="Comma-separated list of prompt versions (default: v1.4.0,v1.0.0,v1.1.0,v1.3.0)")
    parser.add_argument("--data", 
                       default="../data/essay_writing_40_sample.xlsx",
                       help="Path to input Excel file")
    parser.add_argument("--output", 
                       default="batch_evaluation_comparison.xlsx",
                       help="Output Excel file name")
    parser.add_argument("--checkpoint", 
                       default="batch_evaluation_checkpoint.json",
                       help="Checkpoint file for resume capability")
    parser.add_argument("--batch-size", 
                       type=int,
                       default=5,
                       help="Number of API calls before saving checkpoint (default: 5)")
    parser.add_argument("--resume", 
                       action="store_true",
                       help="Resume from existing checkpoint if available")
    
    args = parser.parse_args()
    
    # prompt 버전 파싱
    prompt_versions = [v.strip() for v in args.versions.split(",")]
    
    # 경로 설정
    data_path = args.data
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"comparison_{timestamp}_{args.output}"
    
    # 배치 평가기 초기화 (checkpoint 파일과 함께)
    evaluator = EssayBatchEvaluator(
        prompt_versions=prompt_versions,
        checkpoint_file=args.checkpoint
    )
    evaluator.batch_size = args.batch_size
    
    try:
        logger.info(f"🚀 Starting batch evaluation comparison")
        logger.info(f"📝 Prompt versions: {prompt_versions}")
        logger.info(f"📁 Data file: {data_path}")
        logger.info(f"📁 Output file: {output_path}")
        logger.info(f"💾 Checkpoint file: {args.checkpoint}")
        logger.info(f"📦 Batch size: {args.batch_size}")
        if args.resume:
            logger.info("🔄 Resume mode enabled")
        
        # 1. 샘플 데이터 로드
        logger.info("🔄 Loading sample data...")
        df = evaluator.load_sample_data(data_path)
        
        # 2. 배치 평가 실행
        logger.info("🔄 Starting batch evaluation...")
        results = evaluator.process_all_essays(df)
        
        # 3. 엑셀 보고서 생성
        logger.info("🔄 Creating Excel report...")
        evaluator.create_excel_report(results, output_path)
        
        # 4. Checkpoint 파일 정리 (완료 후)
        if Path(args.checkpoint).exists():
            Path(args.checkpoint).unlink()
            logger.info(f"🗑️ Removed checkpoint file: {args.checkpoint}")
        
        logger.info("🎉 Batch evaluation comparison completed successfully!")
        logger.info(f"📊 Results saved to: {output_path}")
        
    except KeyboardInterrupt:
        logger.info("\n⏹️ Evaluation interrupted by user")
        # 인터럽트 시에도 체크포인트 저장
        if 'evaluator' in locals():
            evaluator.save_checkpoint()
        logger.info(f"💾 Progress saved in checkpoint: {args.checkpoint}")
        logger.info("🔄 Resume with: python excel_creation.py --resume")
    except Exception as e:
        logger.error(f"❌ Batch evaluation failed: {e}")
        logger.info(f"💾 Progress may be saved in checkpoint: {args.checkpoint}")
        raise
        
        logger.info("🎉 Batch evaluation comparison completed successfully!")
        logger.info(f"📊 Results saved to: {output_path}")
        
    except Exception as e:
        logger.error(f"❌ Batch evaluation failed: {e}")
        raise
        raise

if __name__ == "__main__":
    main()