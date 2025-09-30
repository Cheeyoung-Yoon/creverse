# %%
"""
Essay Evaluation Batch Processing Script

이 스크립트는 40개 샘플 에세이를 4개 레벨(Basic, Intermediate, Advanced, Expert)로 
각각 평가하여 총 160개의 결과를 엑셀 파일로 생성합니다.
"""

import asyncio
import json
import logging
import pandas as pd
import requests
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('batch_evaluation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EssayBatchEvaluator:
    """배치 에세이 평가기"""
    
    def __init__(self, api_url: str = "http://localhost:8000/v1/essay-eval"):
        self.api_url = api_url
        self.levels = ["Basic", "Intermediate", "Advanced", "Expert"]
        self.results = {}
        
    def load_sample_data(self, excel_path: str) -> pd.DataFrame:
        """샘플 에세이 데이터 로드"""
        try:
            df = pd.read_excel(excel_path)
            # 첫 번째 행이 헤더인지 확인하고 NaN 값이 있는 행 제거
            df = df.dropna(subset=['submit_text'])
            
            logger.info(f"✅ Loaded {len(df)} essays from {excel_path}")
            logger.info(f"📊 Columns: {list(df.columns)}")
            logger.info(f"📋 Original levels distribution:")
            if 'rubric_level' in df.columns:
                level_counts = df['rubric_level'].value_counts()
                for level, count in level_counts.items():
                    logger.info(f"   {level}: {count} essays")
            
            return df
        except Exception as e:
            logger.error(f"❌ Failed to load Excel file: {e}")
            raise
    
    def call_evaluation_api(self, essay_text: str, topic_prompt: str, level_group: str) -> Dict[str, Any]:
        """API 호출하여 에세이 평가"""
        payload = {
            "level_group": level_group,
            "topic_prompt": topic_prompt,
            "submit_text": essay_text
        }
        
        try:
            response = requests.post(
                self.api_url,
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=60  # 60초 타임아웃
            )
            
            if response.status_code == 200:
                result = response.json()
                logger.debug(f"✅ API call successful for level {level_group}")
                return {
                    "status": "success",
                    "data": result,
                    "response_time": response.elapsed.total_seconds()
                }
            else:
                logger.error(f"❌ API call failed with status {response.status_code}")
                return {
                    "status": "error",
                    "error": f"HTTP {response.status_code}: {response.text}",
                    "response_time": response.elapsed.total_seconds()
                }
                
        except requests.exceptions.Timeout:
            logger.error("❌ API call timed out")
            return {"status": "timeout", "error": "Request timed out"}
        except Exception as e:
            logger.error(f"❌ API call failed: {e}")
            return {"status": "error", "error": str(e)}
    
    def process_all_essays(self, df: pd.DataFrame) -> Dict[str, List[Dict]]:
        """모든 에세이를 모든 레벨로 평가"""
        results = {level: [] for level in self.levels}
        total_calls = len(df) * len(self.levels)
        current_call = 0
        
        logger.info(f"🚀 Starting batch evaluation: {len(df)} essays × {len(self.levels)} levels = {total_calls} API calls")
        
        for level in self.levels:
            logger.info(f"📊 Processing level: {level}")
            
            for idx, row in df.iterrows():
                current_call += 1
                essay_text = str(row.get('submit_text', ''))  # 실제 컬럼명 사용
                essay_id = row.get('essay_id', idx)
                original_level = row.get('rubric_level', 'unknown')
                topic_prompt = row.get('topic_prompt', '')
                
                if not essay_text or essay_text.strip() == '':
                    logger.warning(f"⚠️ Empty essay text at row {idx} (essay_id: {essay_id})")
                    continue
                
                logger.info(f"[{current_call}/{total_calls}] Evaluating essay {essay_id} (original: {original_level}) with level {level}")
                logger.debug(f"📝 Essay preview: {essay_text[:100]}...")
                
                # API 호출
                api_result = self.call_evaluation_api(essay_text, topic_prompt, level)
                
                # 결과 정리
                # 기본 정보 기록
                result_record = {
                    "essay_id": essay_id,
                    "original_level": original_level,
                    "evaluation_level": level,
                    "topic_prompt": topic_prompt,
                    "essay_text": essay_text[:500] + "..." if len(essay_text) > 500 else essay_text,  # 텍스트 길이 제한
                    "essay_length": len(essay_text),
                    "response_time": api_result.get("response_time", 0),
                    "api_status": api_result.get("status", "unknown")
                }
                
                if api_result["status"] == "success":
                    eval_data = api_result["data"]
                    
                    # 전체 평가 정보 추가
                    if "aggregated" in eval_data:
                        agg_data = eval_data["aggregated"]
                        result_record["total_score"] = agg_data.get("score", 0)
                        result_record["total_corrections_count"] = len(agg_data.get("corrections", []))
                    
                    # grammar 섹션 처리
                    if "grammar" in eval_data:
                        grammar = eval_data["grammar"]
                        result_record["grammar_score"] = grammar.get("score", 0)
                        result_record["grammar_feedback"] = grammar.get("feedback", "")[:500]
                        result_record["grammar_corrections_count"] = len(grammar.get("corrections", []))
                        
                        corrections = grammar.get("corrections", [])
                        if corrections:
                            first_correction = corrections[0]
                            result_record["grammar_first_correction"] = f"{first_correction.get('highlight', '')} → {first_correction.get('correction', '')}"[:200]
                    
                    # structure 안의 섹션들 처리
                    if "structure" in eval_data:
                        structure_data = eval_data["structure"]
                        for section_name in ["introduction", "body", "conclusion"]:
                            if section_name in structure_data:
                                section = structure_data[section_name]
                                result_record[f"{section_name}_score"] = section.get("score", 0)
                                result_record[f"{section_name}_feedback"] = section.get("feedback", "")[:500]
                                result_record[f"{section_name}_corrections_count"] = len(section.get("corrections", []))
                                
                                # 첫 번째 correction만 기록
                                corrections = section.get("corrections", [])
                                if corrections:
                                    first_correction = corrections[0]
                                    result_record[f"{section_name}_first_correction"] = f"{first_correction.get('highlight', '')} → {first_correction.get('correction', '')}"[:200]
                    
                    # 타이밍 정보
                    if "timings" in eval_data:
                        result_record["total_processing_time"] = eval_data["timings"].get("total", 0) / 1000  # ms를 초로 변환
                        
                else:
                    result_record["error"] = api_result.get("error", "Unknown error")
                
                results[level].append(result_record)
                
                # API 호출 간격 (서버 부하 방지)
                time.sleep(1)
        
        logger.info("✅ Batch evaluation completed!")
        return results
    
    def create_excel_report(self, results: Dict[str, List[Dict]], output_path: str):
        """결과를 엑셀 파일로 저장"""
        logger.info(f"📝 Creating Excel report: {output_path}")
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for level in self.levels:
                level_results = results[level]
                if not level_results:
                    logger.warning(f"⚠️ No results for level {level}")
                    continue
                
                # DataFrame 생성
                df = pd.DataFrame(level_results)
                
                # 시트에 저장
                df.to_excel(writer, sheet_name=f"{level}_Results", index=False)
                
                # 스타일링
                workbook = writer.book
                worksheet = writer.sheets[f"{level}_Results"]
                
                # 헤더 스타일링
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # 열 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # 최대 50자
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                logger.info(f"✅ Created sheet: {level}_Results with {len(level_results)} records")
            
            # 요약 시트 생성
            self.create_summary_sheet(writer, results)
        
        logger.info(f"✅ Excel report saved: {output_path}")
    
    def create_summary_sheet(self, writer, results: Dict[str, List[Dict]]):
        """요약 시트 생성"""
        summary_data = []
        
        for level in self.levels:
            level_results = results[level]
            if not level_results:
                continue
            
            successful_calls = len([r for r in level_results if r.get("api_status") == "success"])
            failed_calls = len([r for r in level_results if r.get("api_status") != "success"])
            avg_response_time = sum([r.get("response_time", 0) for r in level_results]) / len(level_results)
            
            # 평균 점수 계산 (성공한 호출만)
            successful_results = [r for r in level_results if r.get("api_status") == "success"]
            avg_scores = {}
            
            if successful_results:
                sections = ["introduction", "body", "conclusion", "grammar"]
                for section in sections:
                    scores = [r.get(f"{section}_score", 0) for r in successful_results if f"{section}_score" in r]
                    avg_scores[f"avg_{section}_score"] = sum(scores) / len(scores) if scores else 0
            
            summary_record = {
                "Level": level,
                "Total_Essays": len(level_results),
                "Successful_Calls": successful_calls,
                "Failed_Calls": failed_calls,
                "Success_Rate": f"{(successful_calls/len(level_results)*100):.1f}%" if level_results else "0%",
                "Avg_Response_Time": f"{avg_response_time:.2f}s",
                **avg_scores
            }
            
            summary_data.append(summary_record)
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        
        # 요약 시트 스타일링
        workbook = writer.book
        worksheet = writer.sheets["Summary"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        logger.info("✅ Created Summary sheet")

def main():
    """메인 실행 함수"""
    # 경로 설정
    data_path = "/home/ycy/work_dir/creverse/data/essay_writing_40_sample.xlsx"
    output_path = "/home/ycy/work_dir/creverse/excel_creation.xlsx"
    
    # 배치 평가기 초기화
    evaluator = EssayBatchEvaluator()
    
    try:
        # 1. 샘플 데이터 로드
        logger.info("🔄 Loading sample data...")
        df = evaluator.load_sample_data(data_path)
        
        # 2. 배치 평가 실행
        logger.info("🔄 Starting batch evaluation...")
        results = evaluator.process_all_essays(df)
        
        # 3. 엑셀 보고서 생성
        logger.info("🔄 Creating Excel report...")
        evaluator.create_excel_report(results, output_path)
        
        logger.info("🎉 Batch evaluation completed successfully!")
        logger.info(f"📊 Results saved to: {output_path}")
        
    except Exception as e:
        logger.error(f"❌ Batch evaluation failed: {e}")
        raise

if __name__ == "__main__":
    main()